import numpy as np
import pandas as pd
import openpyxl as xl
import datetime
from shutil import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from isdayoff import ProdCalendar, DayType, date as datee


"""
cell height for the name of the person signing the document = CHFTNOTPSTD = CTD
высота ячейки, для названия должности сотрудника, который будет подписывать документ
данную величину можно регулировать по усмотрению(x пикселей)
------------------------------------------------
jump - через если клетка слита через 4 ячейки, т это значит что после функии нужно идти на 4 вниз, к следующей
start_col - первый столбец, откуда все и начинается
last_col - последний столбец
"""


def unmerge_cells_fun(jump, start_col, last_col):
    row = r
    for row in range(row, LAST_ROW, jump):
        for col in [(start_col, last_col)]:
            sheet.unmerge_cells(start_row=row, start_column=col[0], end_row=row + jump - 1, end_column=col[1])

# все 4 снизу делают почти тоже самое, что и start_col_fun(start_col). Имеет более локальный вид


def cell_function_days_U(cell_num_days: int) -> str:  # добавление отработанных дней в колонку U
    days_cell = str("U") + str(cell_num_days)  # создаем номер ячейки, куда будем заполнять данные
    return days_cell


def cell_function_hours_U(cell_num_hours: int) -> str:  # добавление отработанных часов в колонку U
    hours_cell = str("U") + str(cell_num_hours)
    return hours_cell


def cell_function_days_V(cell_num_days: int) -> str:  # добавление отработанных дней в колонку V
    days_cell_2 = str("V") + str(cell_num_days)  # создаем номер ячейки, куда будем заполнять данные
    return days_cell_2


def cell_function_hours_V(cell_num_hours: int) -> str:  # добавление отработанных часов в колонку V
    hours_cell_2 = str("V") + str(cell_num_hours)
    return hours_cell_2


"""
last_col - последний столбец
fpos_row -первая строка, где начинаем соединять
lpos_row - послдная строка, где заканчиваем соединять
k - некоторые вещи нам нужно соединить один раз(без цикла), и удобнее просто указать их место
"""


def merge_cells_fun(start_col: int, last_col: int, fpos_row: int, lpos_row: int, k: int):
    row = r + k
    for col in [(start_col, last_col)]:
        sheet.merge_cells(start_row=row + fpos_row, start_column=col[0], end_row=row + lpos_row, end_column=col[1])


"""
используется для добавления техта в поля для департмента, дата создания файла, с какой по какой период. Также выравнивание по центру
"""


def add_into_merged_cell(start_col: int, number_of_the_row: int, text):
    top_left_cell = sheet[f'{start_col_fun(start_col)}{number_of_the_row}']  # обозначаем клетку, куда хотим добавить название
    top_left_cell.value = text   # добавляем название
    top_left_cell.alignment = Alignment(horizontal="center", vertical="center")  # выравниваем по цетру
    if (top_left_cell.value is None):
        raise ValueError('В клетку ничего не было записано! Она пустая!')


"""
превращается нумерацию столбцов в буквы, чтобы можно было использовать в коде
Например : 22=> V ; 27 => AA
"""


def start_col_fun(start_col: int) -> str:
    if(start_col == 0):
        raise ValueError('столбец равен нулю! Такого быть не может, ведь нумерация с единицы/первого. Это означает, что ошибка в системе, где-то start_col обнуляется')
    if(start_col > 26):
        start_col -= 26
        start_col_fun_2 = "A" + str(chr(64 + start_col))
    else:
        start_col_fun_2 = str(chr(64 + start_col))
    return start_col_fun_2




"""
start_col - первый столбец, откуда все и начинается
проставляет обычные 5 по 8, с поправкой на ставку. Если выходной (по производственному календарю), то проставляет 0
"""


def standart_5_8_day_fraction(day_in_the_month: int, column_for_work_days: int, day_fraction_for_lazy_ass: int, k: int):
    if calendar.check(datee(2021, 11, day_in_the_month)) == DayType.WORKING:
        sheet.cell(r + k, column_for_work_days).value = "Я"
        sheet.cell(r + k + 1, column_for_work_days).value = 8 * day_fraction_for_lazy_ass
        if (sheet.cell(r + k, column_for_work_days).value == "Я" and sheet.cell(r + k + 1, column_for_work_days).value == 0):
            raise ValueError('Данный день отмечен как рабочий, при этом нет рабочих часов. Скорее всего сбит days_count или row_shift.')
    else:
        sheet.cell(r + k, column_for_work_days).value = "В"
        sheet.cell(r + k + 1, column_for_work_days).value = 0
        if (sheet.cell(r + k, column_for_work_days).value == "В" and sheet.cell(r + k + 1, column_for_work_days).value != 0):
            raise ValueError('Хоть и стоит выходной, в данный день почему-то имеются рабочие часы. Либо опечатка строчкой выше, либо сбит days_count(может что-то еще)')
# ищем два выходных подряд, чтобы туда поставить ночную смену, если она была 15 или 31 числа


def LF_free_space_for_night_shift(column_for_work_days: int, day_fraction_for_lazy_ass: int, k: int):
    for i in range(5, 20):
        if(sheet.cell(r + k, column_for_work_days).value == "В" and sheet.cell(r + k, column_for_work_days + 1).value == "В"):
            lazy_2 = str("{:.0f}".format(2 * day_fraction_for_lazy_ass))
            sheet.cell(r + k, column_for_work_days).value = "Я/Н"
            sheet.cell(r + 1 + k, column_for_work_days).value = lazy_2 + "/" + lazy_2
            sheet.cell(r + k, column_for_work_days + 1).value = "Я/Н"
            sheet.cell(r + 1 + k, column_for_work_days + 1).value = lazy_2 + "/" + str("{:.0f}".format(6 * day_fraction_for_lazy_ass))
            break
        column_for_work_days += 1


'''
все что внизу - наработчки, которые пока не используются
соединяю поля, куда будут вбиваться подписи, и заоодно сраззу эти подписи
пока что не работает
'''


def last_col_fun(last_col: int) -> str:
    if(start_col == 0):
        raise ValueError('столбец равен нулю! Такого быть не может, ведь нумерация с единицы/первого. Это означает, что ошибка в системе, где-то last_col обнуляется')
    last_col_fun_2 = str(chr(64 + last_col))
    return last_col_fun_2


def merge_cells_fun_inside(start_col: int, last_col: int, fpos_row: int, lpos_row: int, k: int, text):
    row = r + k
    for col in [(start_col, last_col)]:
        sheet.merge_cells(start_row=row + fpos_row, start_column=col[0], end_row=row + lpos_row, end_column=col[1])

    add_into_merged_cell(start_col, r + k, text)
